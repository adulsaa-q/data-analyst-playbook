import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# path relative to project root
os.makedirs("phase1_python_basics/section8_ExcelWriter_openpyxl/reports", exist_ok=True)

def generate_report(input_path, output_path):
    """
    Generate an Excel report from cleaned sales data.

    Args:
        input_path (str): Path to the cleaned CSV file
        output_path (str): Path to save the Excel report

    Returns:
        None
    """

    df = pd.read_csv(input_path, encoding='utf-8')

    # convert date
    df["order_date"] = pd.to_datetime(df["order_date"])

    # create month column
    df = df.assign(month=df["order_date"].dt.strftime("%Y-%m"))

    # summary
    total_revenue = df['revenue'].sum()
    total_orders = df['order_id'].nunique()

    overall_summary = {
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'avg_order_value': total_revenue / total_orders,
        'start_date': df['order_date'].min().strftime('%Y-%m-%d'),
        'end_date': df['order_date'].max().strftime('%Y-%m-%d'),
    }

    summary = pd.DataFrame([overall_summary])

    product_revenue = (
        df.groupby("product")
        .agg(total_revenue=("revenue", "sum"))
        .sort_values(by="total_revenue", ascending=False)
    )

    monthly_revenue = (
        df.groupby("month")
        .agg(total_revenue=("revenue", "sum"))
        .sort_index()
    )

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name="Overall Summary", index=False)
        product_revenue.to_excel(writer, sheet_name="Revenue by Product", index=True)
        monthly_revenue.to_excel(writer, sheet_name="Revenue by Month", index=True)

    wb = load_workbook(output_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=True)
    wb.save(output_path)


if __name__ == "__main__":
    # build absolute path relative to this script's location
    base = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(base, "..", "..")

    input_path = os.path.join(project_root, "data", "orders_clean.csv")
    output_path = os.path.join(base, "reports", "summary_report.xlsx")

    generate_report(input_path, output_path)